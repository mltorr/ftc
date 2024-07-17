import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl.styles
from io import BytesIO
import numpy as np
from openpyxl.styles import NamedStyle, PatternFill
from openpyxl.styles import Font
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.utils import get_column_letter


def filter_and_transpose_data(client_name, schedule_no, start_date, end_date):
    # Load the Excel file
    df = pd.read_excel("ftc.xlsx", sheet_name="Dates", parse_dates=["Dates"])

    # Convert start_date and end_date to datetime objects
    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)

    # Convert Dates column to datetime if it's not already
    if df['Dates'].dtype != '<M8[ns]':
        df['Dates'] = pd.to_datetime(df['Dates'])

    # Filter data based on start_date and end_date
    filtered_df = df[(df["Dates"] >= start_date) & (df["Dates"] <= end_date)]

    # Transpose the DataFrame and rename the columns
    transposed_df = filtered_df.set_index("Dates").transpose().reset_index()
    # transposed_df.columns = ["Name"] + [f"Date{i+1}" for i in range(len(transposed_df.columns)-1)]

    return transposed_df

def update_excel_file(name, schedule_no, result_df, combined_table, pivot_df, first_level_column, second_level_column, third_level_column, merge_df, start_date, end_date):
    # Load the workbook
    wb = load_workbook("ftc.xlsx")

    # Select the sheet
    ws1 = wb["(A) FTC Recovery Schedule"]
    ws2 = wb["(B) FTC Sch_working"]
    ws3 = wb["(C) Detail Calculation"]

    # Lock editing for sheets "FTC Recovery Schedule" and "FTC Sch_working"
    for ws in [ws1, ws2]:
        ws.protection = SheetProtection(autoFilter=True, sort=True, sheet=True, objects=True, scenarios=True)

    # Update sheet (A) cells
    ws1["B6"] = name
    ws1["G6"] = f"FTC {schedule_no}"
    ws1["G7"] = datetime.now().strftime("%d.%m.%Y")\

    message = f"1. The above calculations have been prepared based on information provided for the review period {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}."

    # Update cell A66 with the message
    ws1['A66'] = message

    # Update sheet (B) cells
    main2_row = result_df[result_df['index'] == 'Main2']
    main2_values = main2_row.drop(columns=['index']).values.flatten()
    for i, value in enumerate(main2_values, start=15):  # Start from column O (index 15)
        cell = ws3.cell(row=4, column=i)
        cell.value = value

    # Update sheet (B) cells
    mainb_row = result_df[result_df['index'] == 'Main2']
    mainb_values = mainb_row.drop(columns=['index']).values.flatten()
    for i, valb in enumerate(mainb_values, start=5):  # Start from column O (index 15)
        cell = ws2.cell(row=7, column=i)
        cell.value = valb

    # Define the font color
    font_color = 'FFFFFF'  # White color

    # Define named style for date format
    date_style = NamedStyle(name='date_style', number_format='MM/DD/YYYY')

    # Update sheet (B) cells for Rates table
    mainb2_row = result_df[result_df['index'] == 'Column']
    mainb2_values = mainb2_row.drop(columns=['index']).values.flatten()
    for i, valb2 in enumerate(mainb2_values, start=5):  # Start from column O (index 15)
        cell = ws2.cell(row=6, column=i)
        if isinstance(valb2, pd.Timestamp):  # Check if the value is a timestamp
            cell.value = valb2
            cell.style = date_style  # Apply date style
        else:
            cell.value = valb2
        cell.font = Font(color=font_color)  # Set font color to white

    # Update sheet (B) cells vertically in one column
    maina_row = result_df[result_df['index'] == 'Main2']
    maina_values = maina_row.drop(columns=['index']).values.flatten()

    # Start from row 7 (index 7) and iterate over each value
    for i, vala in enumerate(maina_values, start=10):
        cell = ws1.cell(row=i, column=1)  # Column A (index 10)
        cell.value = vala

    # Create a new DataFrame with only the grouping columns of pivot_df
    selected_cols = []
    if first_level_column != 'None' and first_level_column in pivot_df.columns:
        selected_cols.append(first_level_column)
    if second_level_column != 'None' and second_level_column in pivot_df.columns:
        selected_cols.append(second_level_column)
    if third_level_column != 'None' and third_level_column in pivot_df.columns:
        selected_cols.append(third_level_column)

    # Write the column names to cells B4, C4, and D4
    for i, col in enumerate(selected_cols, start=2):
        cell = ws3.cell(row=4, column=i)
        cell.value = col
    
    drop_df = combined_table.drop(selected_cols, axis=1)

    # Write combined_table values to sheet C starting at cell O5
    for r_idx, row in enumerate(drop_df.iterrows(), start=5):
        for c_idx, val in enumerate(row[1]):
            cell = ws3.cell(row=r_idx, column=c_idx + 15)  # Start from column O (index 15)
            cell.value = val
            cell.number_format = '#,##0.00_);[Black](#,##0.00)'

    # Calculate totals per column and add them to the last row
    for c_idx in range(15, len(drop_df.columns) + 15):
        total_formula = f"SUM({ws3.cell(row=5, column=c_idx).coordinate}:{ws3.cell(row=r_idx, column=c_idx).coordinate})"
        total_cell = ws3.cell(row=r_idx + 1, column=c_idx)  # Add to the row below the last data row
        total_cell.value = f"={total_formula}"
        total_cell.number_format = '#,##0.00_);[Black](#,##0.00)'
        total_cell.font = Font(bold=True)

    last_row_index = len(pivot_df) + 5

    # Calculate the sum formula for column N starting from the last row index
    sum_formula = f"SUM(O{last_row_index}:BS{last_row_index})"

    # Apply the sum formula to the cell in column N in the row after the last row
    cell = ws3.cell(row=last_row_index, column=14)  # Column N
    cell.value = f"={sum_formula}"

    # Set the font to bold for the cell
    cell.font = Font(bold=True)

    # Create a style for the percentage format
    percent_style = NamedStyle(name='percent')
    percent_style.number_format = '0.00%'

    # Apply the percentage format to the specified columns
    for col in ['I', 'J', 'K', 'L']:
        for row in range(5, len(drop_df) + 5):
            cell = ws3[col + str(row)]  # Get the cell
            cell.style = percent_style  # Apply the style

            # Apply grey background color
            grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.fill = grey_fill

    # Write the values of the grouping columns of pivot_df to Sheet C starting from cell B5
    for r_idx, (_, row) in enumerate(merge_df.iterrows(), start=5):
        for c_idx, val in enumerate(row):
            cell = ws3.cell(row=r_idx, column=c_idx + 2)  # Start from column B (index 2)
            cell.value = val

    sum_formula_template = f"SUM('{ws3.title}'!O{{row}}:BS{{row}})"

    # Assigning formulas to cells N5 and below
    for i in range(5, len(drop_df) + 5):
        sum_formula = sum_formula_template.format(row=i)
        cell = ws3.cell(row=i, column=14)  # Column N
        cell.value = f"={sum_formula}"

    # Remove the "Dates" sheet
    if "Dates" in wb.sheetnames:
        wb.remove(wb["Dates"])

    # Save the workbook
    output = BytesIO()
    wb.save(output)

    return output.getvalue()



def pivot_table(df, date_column, litres_column, first_level_column, second_level_column=None, third_level_column=None):
    # Create a copy of the DataFrame
    pivot_df = df.copy()

    # Pivot the DataFrame
    pivot_df = pivot_df.pivot_table(index=[first_level_column, second_level_column, third_level_column], columns=date_column, values=litres_column)

    return pivot_df

def main():
    st.title("FTC Automation v0.1")

    # Calculate default start date (end_date - 50 months)
    end_date_default = datetime.now().replace(day=1)
    start_date_default = end_date_default - timedelta(days=48*30)  # Assuming 30 days per month
    start_date_default = start_date_default.replace(day=1)  # Setting day to 1

    # Calculate default end date (end of last month)
    end_date_default = datetime.now().replace(day=1) - timedelta(days=1)

    # User inputs
    client_name = st.text_input("Enter Client Name (and ABN):")
    schedule_no = st.text_input("Enter Schedule No.:")
    start_date = st.date_input("Enter Start Period:", value=start_date_default)
    end_date = st.date_input("Enter End Period:", value=end_date_default)

    # Upload Excel file
    uploaded_file = st.file_uploader("Upload Fuel Consumption Data", type=["xlsx"])
    if uploaded_file:
        # Ask for sheet name
        # sheet_name = st.text_input("Enter Sheet Name:")

        # Read Excel file and extract selected sheet
        df = None
        if uploaded_file:
            try:
                # Get the list of sheet names
                excel_sheets = pd.ExcelFile(uploaded_file).sheet_names

                # Add a default option for None
                excel_sheets.insert(0, "None")

                # Ask the user to select a sheet from the available options
                sheet_name = st.selectbox("Select Sheet Name:", options=excel_sheets)

                if sheet_name != "None":
                    # Read the selected sheet
                    df = pd.read_excel(uploaded_file, sheet_name=sheet_name)

                    # Display the uploaded table
                    st.write("Uploaded Table:")
                    st.write(df)
                else:
                    st.warning("Please select a sheet.")
            except Exception as e:
                st.error(f"Error: {e}")

        # Upload Excel file for the description table
        uploaded_description_file = st.file_uploader("Upload Vehicle Data (Excel)", type=["xlsx"])
        description_df = None
        description_sheet_name = None
        joining_key = None

        if uploaded_description_file:
            try:
                # Get the list of sheet names
                description_excel_sheets = pd.ExcelFile(uploaded_description_file).sheet_names

                # Add a default option for None
                description_excel_sheets.insert(0, "None")

                # Ask the user to select a sheet from the available options
                description_sheet_name = st.selectbox("Select Description Table Sheet Name:", options=description_excel_sheets)

                if description_sheet_name != "None":
                    # Read the selected sheet
                    description_df = pd.read_excel(uploaded_description_file, sheet_name=description_sheet_name)

                    # Display the uploaded description table
                    st.write("Uploaded Vehicle Data:")
                    st.write(description_df)

                    # Ask for joining key
                    if description_df is not None:
                        joining_key = st.selectbox("Select Joining Key (for Merging the Fuel Consumption with the Vehicle Data ):", options=description_df.columns)
                else:
                    st.warning("Please select a sheet for the description table.")
            except Exception as e:
                st.error(f"Error: {e}")

        # Ask for column selections
        if df is not None:
            date_column = st.selectbox("Select Date Column for FTC", options=df.columns)
            litres_column = st.selectbox("Select Fuel Consumption Column", options=df.columns)
            first_level_column = st.selectbox("Select First Level Column Grouping", options=df.columns)
            second_level_column = st.selectbox("Select Second Level Column Grouping", options=['None'] + list(df.columns))
            third_level_column = st.selectbox("Select Third Level Column Grouping", options=['None'] + list(df.columns))

            # Button to trigger the data processing
            if st.button("Process Data"):
                # Check if all fields are filled
                if date_column and litres_column and first_level_column:
                    
                    # pivot_df = pivot_table(df, date_column, litres_column, first_level_column, second_level_column, third_level_column)
                    if not isinstance(start_date, pd.Timestamp):
                        start_date = pd.to_datetime(start_date)
                    if not isinstance(end_date, pd.Timestamp):
                        end_date = pd.to_datetime(end_date)
                    df_filtered = df[df[date_column].between(start_date, end_date)]
                    pivot_df = df_filtered.pivot_table(index=[first_level_column, second_level_column, third_level_column], 
                              columns=date_column, 
                              values=litres_column, 
                              aggfunc='sum')
                    pivot_df.reset_index(inplace=True)
                    # pivot_df.to_excel("pivot_df.xlsx", index=True)
                    # pivot_df_imported = pd.read_excel("pivot_df.xlsx", index_col=0)
                    # selected_columns = [col for col in pivot_df_imported.columns if col in [first_level_column, second_level_column, third_level_column]]
                    # selected_df = pivot_df_imported[selected_columns]

                    # st.write("Pivoted Table:")
                    # st.write(pivot_df)
                    # st.write(selected_df)

                    # Create a new table from result_df without the "index" column
                    result_df = filter_and_transpose_data(client_name, schedule_no, start_date, end_date)
                    new_table = result_df.drop(columns=['index'])
                    new_table = new_table.drop(new_table.index)
                    new_table = new_table.astype(float)

                    new_table[first_level_column] = np.nan
                    new_table[second_level_column] = np.nan
                    new_table[third_level_column] = np.nan

                    # Get a list of all column names except the ones you just added
                    existing_columns = new_table.columns.difference([first_level_column, second_level_column, third_level_column])

                    # Reorder the columns so that the new columns appear first, followed by the existing columns
                    new_table = new_table[[first_level_column, second_level_column, third_level_column, *existing_columns]]

                    # st.write(new_table)
                    
                    combined_table = pd.concat([pivot_df, new_table], ignore_index=True)
                    combined_table.columns = combined_table.columns.astype(str)
                    combined_table = combined_table[sorted(combined_table.columns)]
                    existing_columns = combined_table.columns.difference([first_level_column, second_level_column, third_level_column])
                    combined_table = combined_table[[first_level_column, second_level_column, third_level_column, *existing_columns]]

                    # Drop the index for new_table and remove the column "index"
                    # df_sort = combined_table.sort_index(axis=1)

                    st.write("Pivot Table:")
                    st.write(combined_table)

                    if combined_table is not None and description_df is not None and joining_key is not None:
                        # Select columns from combined_table to merge
                        merge_cols = [first_level_column, second_level_column, third_level_column]

                        # Merge the tables on the joining key
                        merge_df = pd.merge(combined_table[merge_cols], description_df, on=joining_key, how='left')

                        # Display the merged table
                        st.write("Merged Vehicle Data:")
                        st.write(merge_df)

                        # st.write(result_df)
                    
                    # st.write("Dates Table:")
                    # st.write(df_sort)
                    
                    excel_bytes = update_excel_file(client_name, schedule_no, result_df, combined_table, pivot_df, first_level_column, second_level_column, third_level_column, merge_df, start_date, end_date)
                    st.download_button("Download Excel", data=excel_bytes, file_name="FTC Report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                if start_date > end_date:
                    st.error("End Date must be greater than or equal to Start Date.")
                    return
                else:
                    st.success("Processing has been completed!")

if __name__ == "__main__":
    main()
